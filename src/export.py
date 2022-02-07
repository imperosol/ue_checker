import openpyxl.styles.fonts
from openpyxl import Workbook
from .custom_types import ue_set
import json


def __get_max_sub_elem(content: ue_set) -> dict[str, int]:
    """
    take in argument a nested dictionary
    and return a dictionary which associates each key of inner
    dictionary with the max number of elements in field.
    Example : ::
        {
            'TC1': {'CS': [1, 1], 'TM': [1],       'ME': [1]},
            'TC2': {'CS': [1, 1], 'TM': [1, 1, 1], 'ME': [1, 1]}
        }

    as an input will make the function
    return : ::
        {'CS': 2, 'TM': 3, 'ME': 2}
    """
    max_sizes = dict()
    for elem in content.values():
        for sub_title, sub_elem in elem.items():
            nb_rows = len(sub_elem)
            if sub_title in max_sizes:
                max_sizes[sub_title] = max(max_sizes[sub_title], nb_rows)
            else:
                max_sizes[sub_title] = nb_rows
    return max_sizes


def __get_cum_row_per_cat(row_per_category: dict[str, int]) -> tuple[int, dict[str, int]]:
    """
    Build and return a cumulative dict based on the dict given and the value of the sum of everything.
    Example: ::
        {'a': 1, 'b': 5, 'c': 3, 'd': 4}
    Will
    return: ::
        tuple(13, {'a': 1, 'b': 6, 'c': 9, 'd': 13}
    """
    cum_sum = 0
    result_dict = dict()
    for key, nb_rows in row_per_category.items():
        result_dict[key] = nb_rows + cum_sum
        cum_sum += nb_rows
    return cum_sum, result_dict


def __get_content_rows(content: ue_set, row_per_category: dict[str, int] = None) -> list[list[list[str, str, str]]]:
    """
    Build a list of list of elements corresponding to the content of the table to export from a nested dict
    :param content: a nested dict with the content to export
    :param row_per_category: a dictionary association each category of ue to the number of rows to display it.
    If it is not given, the function builds it.
    :return: a nested list corresponding to the content to display. Each cell will be set like [name, letter, credits].
    Cells which must be empty will have the value ['', '', ''].
    """
    if row_per_category is None:
        row_per_category = __get_max_sub_elem(content)
    total_rows, cumulative_rows = __get_cum_row_per_cat(row_per_category)
    result = [[['', '', ''] for _ in range(len(content))] for _ in range(total_rows)]
    for col, elem in enumerate(content.values()):
        for category, sub_elem in elem.items():
            cat_starting_row = cumulative_rows[category] - row_per_category[category]
            for row, ue in enumerate(sub_elem):
                if len(ue) == 3:
                    result[cat_starting_row + row][col] = ue
                else:
                    for i, info in enumerate(ue):
                        result[cat_starting_row + row][col][i] = info
    return result


def to_xls(content: ue_set) -> str:
    """ Function to create an Excel file with the datas of the student file.
    :return: the name of the file created
    :rtype: str
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ues"
    font = openpyxl.styles.Font(bold=True)
    align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    row_per_category = __get_max_sub_elem(content)
    table = __get_content_rows(content, row_per_category)
    for col, col_title in enumerate(content.keys()):
        worksheet.merge_cells(start_row=1, start_column=2 + col * 3, end_row=1, end_column=4 + col * 3)
        worksheet.cell(1, 2 + col * 3).value = col_title
        worksheet.cell(1, 2 + col * 3).font = font
        worksheet.cell(1, 2 + col * 3).alignment = align
    sub_key = next(iter(content))
    row = 2
    for row_title in content[sub_key].keys():
        add_row = row_per_category[row_title]
        if add_row > 0:
            worksheet.merge_cells(start_row=row, start_column=1, end_row=row + add_row - 1, end_column=1)
        worksheet.cell(row, 1).value = row_title
        worksheet.cell(row, 1).alignment = align
        row += add_row
    for i, row in enumerate(table):
        j = 2
        for col_group in row:
            for col in col_group:
                worksheet.cell(i + 2, j).value = col
                j += 1
    workbook.save('ues.xlsx')
    return 'ues.xlsx'


def to_json(ues: ue_set) -> str:
    """
    create a json file with the data of the student file.
    :return: the name of the file created
    :rtype: str
    """
    with open('ues.json', "w", encoding='utf-8') as f:
        json.dump(ues, f, ensure_ascii=False, indent=2)
    return 'ues.json'


def to_latex(ues: ue_set) -> str:
    """
    export the data of the student file to LaTeX format
    """
    row_per_category = __get_max_sub_elem(ues)
    categories = iter(row_per_category.keys())
    table = __get_content_rows(ues, row_per_category)
    result = "\\documentclass{article}\n\\usepackage{multirow}\n\\begin{document}\n"
    result += "\\begin{tabular}{|c" + "|ccc" * len(table[0]) + "|}\n\\hline\n"
    result += '\t' + ' & '.join([''] + ["\\multicolumn{3}{c|}{" + title + "}" for title in ues]) + "\\\\\n\\hline\n"
    nb_rows = 0
    current_category = next(categories)
    for row in table:
        if nb_rows == 0:
            result += '\t\\multirow{' + f'{row_per_category[current_category]}' + '}{*}{' + current_category + '}'
        result += '\t' + ' & '.join([''] + [col for item in row for col in item])  # flatten and stringify list
        result += '\\\\\n'
        nb_rows += 1
        if row_per_category[current_category] == nb_rows:
            nb_rows = 0
            result += "\t\\hline\n"
            while True:
                try:
                    current_category = next(categories)
                    if row_per_category[current_category] > 0:
                        break
                except StopIteration:
                    break
    result += '\\end{tabular}\n\\end{document}\n'
    with open("ues.tex", "w") as f:
        f.write(result)
    return "ues.tex"


def to_tex(ues: ue_set) -> str:
    """
    export the data of the student file to LaTeX format
    """
    return to_latex(ues)


def to_html(ues: ue_set):
    """
    export the data of the student file to an HTML table.
    """
    row_per_cat = __get_max_sub_elem(ues)
    table = __get_content_rows(ues, row_per_cat)
    result = "<!DOCTYPE html>\n<html>\n<head>\n<style>table{border-collapse: collapse;}\n" \
             "td,th{\nborder: 1px solid black;\ntext-align: center;\npadding: 10px;}</style>\n</head>\n<body> " \
             "<table>\n  <thead>\n    <tr>\n      <th></th>\n      <th>"
    result += "</th>\n      <th>".join(ues.keys()) + "</th>\n    </tr>\n  </thead>\n  <tbody>\n  "
    rows_to_add = iter(row_per_cat.items())
    rows_in_cat = 0
    for row in table:
        result += "  <tr>\n      "
        while rows_in_cat == 0:
            try:
                next_item = next(rows_to_add)
                cat_name, rows_in_cat = next_item
                if rows_in_cat > 0:
                    result += f"<td rowspan=\"{rows_in_cat}\">{cat_name}</td>\n      "
            except StopIteration:
                break
        result += '<td>' + '</td>\n      <td>'.join("&nbsp;".join(col) if col[0] != '' else '' for col in row)
        result += '</td>\n    </tr>\n  '
        rows_in_cat -= 1
    result += "</tbody>\n</table>\n</body></html>"
    with open('ues.html', 'w') as f:
        f.write(result)
    return 'ues.html'
