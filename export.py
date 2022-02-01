from itertools import accumulate

import openpyxl.styles.fonts
from openpyxl import Workbook
from custom_types import ue_set
import json


def __get_max_sub_elem(content: ue_set) -> dict[str, int]:
    """
    take in argument a nested dictionary
    and return a dictionary which associates each key of inner
    dictionary with the max number of elements in field.
    Example : ::
        {
            'TC1': {'CS': [1, 1], 'TM': [1],       'ME': [1]},
            'TC2': {'CS': [1, 1], 'TM': [1, 1, 1], 'ME': [1, 1]}
        }

    as an input will make the function
    return : ::
            {'CS': 2, 'TM': 3, 'ME': 2}
    """
    max_sizes = dict()
    for elem in content.values():
        for sub_title, sub_elem in elem.items():
            nb_rows = len(sub_elem)
            if sub_title in max_sizes:
                max_sizes[sub_title] = max(max_sizes[sub_title], nb_rows)
            else:
                max_sizes[sub_title] = nb_rows
    return max_sizes


def __get_cum_row_per_cat(row_per_category: dict[str, int]) -> tuple[int, dict[str, int]]:
    cum_sum = 0
    result_dict = dict()
    for key, nb_rows in row_per_category.items():
        result_dict[key] = nb_rows + cum_sum
        cum_sum += nb_rows
    return cum_sum, result_dict


def __get_content_rows(content: ue_set, row_per_category: dict[str, int] = None) -> list[list[list[str, str, str]]]:
    if row_per_category is None:
        row_per_category = __get_max_sub_elem(content)
    total_rows, cumulative_rows = __get_cum_row_per_cat(row_per_category)
    result = [[['', '', ''] for _ in range(len(content))] for _ in range(total_rows)]
    for col, elem in enumerate(content.values()):
        for category, sub_elem in elem.items():
            cat_starting_row = cumulative_rows[category] - row_per_category[category]
            for row, ue in enumerate(sub_elem):
                if len(ue) == 3:
                    result[cat_starting_row + row][col] = ue
                else:
                    for i, info in enumerate(ue):
                        result[cat_starting_row + row][col][i] = info
    return result


def to_xls(content: ue_set) -> str:
    """ Function to create an Excel file with the datas of the student file.
    :return: the name of the file created
    :rtype: str
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ues"
    font = openpyxl.styles.Font(bold=True)
    align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    row_per_category = __get_max_sub_elem(content)
    col = 2
    for title, elem in content.items():
        worksheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        worksheet.cell(1, col).value = title
        worksheet.cell(1, col).font = font
        worksheet.cell(1, col).alignment = align
        row = 2
        for sub_title, sub_elem in elem.items():
            size = 0
            add_row = row_per_category[sub_title]
            if add_row > 0:
                worksheet.merge_cells(start_row=row, start_column=1, end_row=row + add_row - 1, end_column=1)
            worksheet.cell(row, 1).value = sub_title
            worksheet.cell(row, 1).alignment = align
            if len(sub_elem) > 0:
                for ue in sub_elem:
                    if len(ue) == 0:
                        continue
                    for i, ue_info in enumerate(ue):
                        if ue_info.isdigit():
                            ue_info = int(ue_info)
                        worksheet.cell(row + size, col + i).value = ue_info
                    size += 1
            row += add_row
        col += 3
    workbook.save('ues.xlsx')
    return 'ues.xlsx'


def to_json(ues: ue_set) -> str:
    """
    create a json file with the datas of the student file.
    :return: the name of the file created
    :rtype: str
    """
    with open('ues.json', "w", encoding='utf-8') as f:
        json.dump(ues, f, ensure_ascii=False, indent=2)
    return 'ues.json'


def to_latex(ues: ue_set) -> str:
    row_per_category = __get_max_sub_elem(ues)
    categories = iter(row_per_category.keys())
    table = __get_content_rows(ues, row_per_category)
    result = "\\documentclass{article}\n\\usepackage{multirow}\n\\begin{document}\n"
    result += "\\begin{tabular}{|c" + "|ccc" * len(table[0]) + "|}\n\\hline\n"
    result += '\t' + ' & '.join([''] + ["\\multicolumn{3}{c|}{" + title + "}" for title in ues]) + "\\\\\n\\hline\n"
    nb_rows = 0
    current_category = next(categories)
    for row in table:
        if nb_rows == 0:
            result += '\t\\multirow{' + f'{row_per_category[current_category]}' + '}{*}{' + current_category + '}'
        result += '\t' + ' & '.join([''] + [col for item in row for col in item])  # flatten and stringify list
        result += '\\\\\n'
        nb_rows += 1
        if row_per_category[current_category] == nb_rows:
            nb_rows = 0
            result += "\t\\hline\n"
            while True:
                try:
                    current_category = next(categories)
                    if row_per_category[current_category] > 0:
                        break
                except StopIteration:
                    break
    result += '\\end{tabular}\n\\end{document}\n'
    with open("ues.tex", "w") as f:
        f.write(result)
    return "ues.tex"


def to_tex(ues: ue_set) -> str:
    return to_latex(ues)
